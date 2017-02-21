# -*- coding: utf-8 -*-
"""
Created on Thu Sep 08 15:09:49 2016

@author: tih
"""
import numpy as np
import os
import gdal
from math import sin, cos, pi
import re
import subprocess
from openpyxl import load_workbook
import time
import osr
import shutil
import datetime
from pyproj import Proj, transform

def main():
############################## INPUT ##########################################		
    # Input for preSEBAL.py
    number = 3                                                                                     # Number defines the column of the inputExcel
    inputExcel=r'D:\Water_Accounting\SEBAL\Python\Newest SEBAL\InputEXCEL_v3_3_6.xlsx'               # The excel with all the SEBAL input data
    VegetationExcel = r'D:\Water_Accounting\SEBAL\Python\Newest SEBAL\Vegetation height model.xlsx'  # This excel defines the p and c factor and vegetation height.
    output_folder = r'D:\Water_Accounting\SEBAL\Python\SEBAL_TEST_DATA\Preprocessing_Output'         # Output folder
    LU_data_FileName=r'D:\Water_Accounting\SEBAL\Tadla_full_scene\Landcover_tadla\LU_map.tif'        # Path to Land Use map

######################## Load Excels ##########################################	
    # Open Excel workbook for SEBAL inputs
    wb = load_workbook(inputExcel)
				
    # Open Excel workbook used for Vegetation c and p factor conversions				
    wb_veg = load_workbook(VegetationExcel, data_only=True)			

######################## Open General info from SEBAL Excel ################### 

    # Open the General_Input sheet			
    ws = wb['General_Input']
			
    # Extract the input and output folder, and Image type from the excel file			
    input_folder = str(ws['B%d' % number].value)                             
    Image_Type = int(ws['D%d' % number].value)                               # Type of Image (1=Landsat & 2 = VIIRS & GLOBA-V)     
    output_folder = os.path.join(input_folder,'Preprocessing_output')
    output_folder_temp = os.path.join(input_folder,'Preprocessing_output','Temp')

    # Create or empty output folder		
    if os.path.isdir(output_folder):
        shutil.rmtree(output_folder)
    os.makedirs(output_folder)	
			
    # Extract the Path to the DEM map from the excel file
    DEM_fileName = '%s' %str(ws['E%d' % number].value) #'DEM_HydroShed_m'  

######################## Extract general data for Landsat ##########################################	
    if Image_Type == 1:	
       
        # Open the Landsat_Input sheet				
        ws = wb['Landsat_Input']		
      
        # Extract Landsat name, number and amount of thermal bands from excel file 
        Name_Landsat_Image = str(ws['B%d' % number].value)    # From glovis.usgs.gov
        Landsat_nr = int(ws['C%d' % number].value)            # Type of Landsat (LS) image used (LS5, LS7, or LS8)
        Bands_thermal = int(ws['D%d' %number].value)         # Number of LS bands to use to retrieve land surface 
 
        # Pixel size of the model
        pixel_spacing=int(30) 

        # the path to the MTL file of landsat				
        Landsat_meta_fileName = os.path.join(input_folder, '%s_MTL.txt' % Name_Landsat_Image)

        # read out the general info out of the MTL file
        year, DOY, hour, minutes, UTM_Zone, Sun_elevation = info_general_metadata(Landsat_meta_fileName) # call definition info_general_metadata
        date=datetime.datetime.strptime('%s %s'%(year,DOY), '%Y %j')
        month = date.month
        day = date.day
######################## Extract general data for VIIRS-PROBAV ##########################################	

    if Image_Type == 2:	

        # Open the VIIRS_PROBAV_Input sheet					
        ws = wb['VIIRS_PROBAV_Input']
    
        # Extract the name of the thermal and quality VIIRS image from the excel file	
        Name_VIIRS_Image_TB = '%s' %str(ws['B%d' % number].value)
      					
        # Extract the name to the PROBA-V image from the excel file	
        Name_PROBAV_Image = '%s' %str(ws['D%d' % number].value)    # Must be a tiff file 
             
        # Pixel size of the model
        pixel_spacing=int(100) 	
 		
        # UTM Zone of the end results					
        UTM_Zone = float(ws['G%d' % number].value)

        #Get time from the VIIRS dataset name (IMPORTANT TO KEEP THE TEMPLATE OF THE VIIRS NAME CORRECT example: npp_viirs_i05_20150701_124752_wgs84_fit.tif)
        Total_Day_VIIRS = Name_VIIRS_Image_TB.split('_')[3]
        Total_Time_VIIRS = Name_VIIRS_Image_TB.split('_')[4]
				
        # Get the information out of the VIIRS name
        year = int(Total_Day_VIIRS[0:4])
        month = int(Total_Day_VIIRS[4:6])
        day = int(Total_Day_VIIRS[6:8])
        Startdate = '%d-%02d-%02d' % (year,month,day)
        DOY=datetime.datetime.strptime(Startdate,'%Y-%m-%d').timetuple().tm_yday
        hour = int(Total_Time_VIIRS[0:2])
        minutes = int(Total_Time_VIIRS[2:4])								

######################## Extract general data from DEM file and create Slope map ##########################################	
    # Variable date name
    Var_name = '%d%02d%02d' %(year, month, day)
								
    # Open DEM and create Latitude and longitude files
    lat,lon,lat_fileName,lon_fileName=DEM_lat_lon(DEM_fileName,output_folder_temp)
    
    # Reproject from Geog Coord Syst to UTM -
    # 1) DEM - Original DEM coordinates is Geographic: lat, lon
    dest, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset(
                   DEM_fileName, pixel_spacing, UTM_Zone=UTM_Zone)
    band = dest.GetRasterBand(1)   # Get the reprojected dem band
    ncol = dest.RasterXSize        # Get the reprojected dem column size
    nrow = dest.RasterYSize        # Get the reprojected dem row size
    shape=[ncol, nrow]
       
    # Read out the DEM band and print the DEM properties
    data_DEM = band.ReadAsArray(0, 0, ncol, nrow)

    # 2) Latitude file - reprojection    
    # reproject latitude to the landsat projection and save as tiff file																																
    lat_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset(
                    lat_fileName, pixel_spacing,  UTM_Zone=UTM_Zone)
 
    # Get the reprojected latitude data															
    lat_proy = lat_rep.GetRasterBand(1).ReadAsArray(0, 0, ncol, nrow)
     
    # 3) Longitude file - reprojection
    # reproject longitude to the landsat projection	 and save as tiff file	
    lon_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset(lon_fileName, pixel_spacing, UTM_Zone=UTM_Zone)

    # Get the reprojected longitude data	
    lon_proy = lon_rep.GetRasterBand(1).ReadAsArray(0, 0, ncol, nrow)
    lon_fileName = os.path.join(output_folder_temp,'lon_resh.tif')
    save_GeoTiff_proy(dest, lon_proy, lon_fileName, shape, nband=1)	
				
    # Calculate slope and aspect from the reprojected DEM
    deg2rad,rad2deg,slope,aspect=Calc_Gradient(data_DEM,pixel_spacing)

    # calculate the coz zenith angle
    dr,coz_zn = Calc_Ra_Mountain(lon,DOY,hour,minutes,lon_proy,lat_proy,slope,aspect) 
    cos_zn_fileName = os.path.join(output_folder_temp,'coz_zn.tif')
    save_GeoTiff_proy(dest, coz_zn, cos_zn_fileName, shape, nband=1)			

    # Reproject coz zenith angle    
    dst_FileName_cos = os.path.join(output_folder_temp,'resh_cos_zn.tif')
    cos_zn_resh = reshape(cos_zn_fileName, dst_FileName_cos, 30, var='cos_zn')

#################### Calculate NDVI and SAVI for LANDSAT ##########################################	

    if Image_Type == 1:	

        # Define bands used for each Landsat number
        if Landsat_nr == 5 or Landsat_nr == 7:
            Bands = np.array([1, 2, 3, 4, 5, 7, 6])
        elif Landsat_nr == 8:
            Bands = np.array([2, 3, 4, 5, 6, 7, 10, 11])  
        else:
            print 'Landsat image not supported, use Landsat 7 or 8'

        # Open MTL landsat and get the correction parameters   
        Landsat_meta_fileName = os.path.join(input_folder, '%s_MTL.txt' % Name_Landsat_Image)
        Lmin, Lmax, k1_c, k2_c = info_band_metadata(Landsat_meta_fileName, Bands)
           
        # Mean solar exo-atmospheric irradiance for each band (W/m2/microm)
        # for the different Landsat images (L5, L7, or L8)
        ESUN_L5 = np.array([1983, 1796, 1536, 1031, 220, 83.44])
        ESUN_L7 = np.array([1997, 1812, 1533, 1039, 230.8, 84.9])
        ESUN_L8 = np.array([1973.28, 1842.68, 1565.17, 963.69, 245, 82.106])
    
        # Open one band - To get the metadata of the landsat images only once (to get the extend)
        src_FileName = os.path.join(input_folder, '%s_B2.TIF' % Name_Landsat_Image)  # before 10!
        ls,band_data,ulx,uly,lrx,lry,x_size_ls,y_size_ls = Get_Extend_Landsat(src_FileName)
         
        # Crop the Landsat images to the DEM extent -
        dst_FileName = os.path.join(output_folder_temp,'cropped_LS_b2.tif')  # Before 10 !!
     							
        # Clip the landsat image to match the DEM map											
        fullCmd = ' '.join(['gdalwarp -te %s %s %s %s' % (ulx_dem, lry_dem,lrx_dem, uly_dem), src_FileName, dst_FileName])
        process = subprocess.Popen(fullCmd)
        process.wait()
   
        # Get the extend of the remaining landsat file	after clipping based on the DEM file	
        lsc,band_data,ulx,uly,lrx,lry,x_size_lsc,y_size_lsc = Get_Extend_Landsat(dst_FileName)
        shape=[x_size_lsc,y_size_lsc]
 			
        # Create the corrected signals of Landsat in 1 array
        Reflect = Landsat_Reflect(Bands,input_folder,Name_Landsat_Image,output_folder_temp,ulx_dem,lry_dem,lrx_dem,uly_dem,shape,Lmax,Lmin,ESUN_L5,ESUN_L7,ESUN_L8,cos_zn_resh,dr,Landsat_nr)

        # Calculate temporal water mask
        water_mask_temp=Water_Mask(shape,Reflect)       

        # Calculate the NDVI and SAVI								
        NDVI,SAVI,albedo = Calc_NDVI_SAVI_albedo(Reflect)

        NDVI_FileName = os.path.join(output_folder,'NDVI_LS_%s.tif'%Var_name)
        save_GeoTiff_proy(dest, NDVI, NDVI_FileName, shape, nband=1)
								
        SAVI_FileName = os.path.join(output_folder,'SAVI_LS_%s.tif'%Var_name)
        save_GeoTiff_proy(dest, SAVI, SAVI_FileName, shape, nband=1)
								
        albedo_FileName = os.path.join(output_folder,'Albedo_LS_%s.tif'%Var_name)
        save_GeoTiff_proy(dest, albedo, albedo_FileName, shape, nband=1)

################### Extract Meteo data for Landsat days from SEBAL Excel ##################
        # Open the Meteo_Input sheet	
        ws = wb['Meteo_Input']	
        # ---------------------------- Instantaneous Air Temperature ------------
        # Open meteo data, first try to open as value, otherwise as string (path)	  
        try:
            Temp_inst = float(ws['B%d' %number].value)                # Instantaneous Air Temperature (°C)

        # if the data is not a value, than open as a string	
        except:
            Temp_inst_name = '%s' %str(ws['B%d' %number].value) 
            Temp_inst_fileName = os.path.join(output_folder, 'Temp', 'Temp_inst_input.tif')
            Temp_inst = Reshape_Reproject_Input_data(Temp_inst_name, Temp_inst_fileName, cos_zn_fileName)

        try:
            RH_inst = float(ws['D%d' %number].value)                # Instantaneous Relative humidity (%)
 
        # if the data is not a value, than open as a string							
        except:
             RH_inst_name = '%s' %str(ws['D%d' %number].value) 
             RH_inst_fileName = os.path.join(output_folder, 'Temp', 'RH_inst_input.tif')
             RH_inst = Reshape_Reproject_Input_data(RH_inst_name, RH_inst_fileName, cos_zn_fileName)	#################### Calculate thermal for Landsat ##########################################			

        Rp = 0.91                        # Path radiance in the 10.4-12.5 µm band (W/m2/sr/µm)
        tau_sky = 0.866                  # Narrow band transmissivity of air, range: [10.4-12.5 µm]
        surf_temp_offset = 3             # Surface temperature offset for water 
 
        esat_inst = 0.6108 * np.exp(17.27 * Temp_inst / (Temp_inst + 237.3)) 													
        eact_inst = RH_inst * esat_inst / 100
        FPAR,tir_emis,Nitrogen,vegt_cover,LAI,b10_emissivity = Calc_vegt_para(NDVI,SAVI,water_mask_temp,shape)

        LAI_FileName = os.path.join(output_folder,'LAI_LS_%s.tif' %Var_name)
        save_GeoTiff_proy(dest, LAI, LAI_FileName, shape, nband=1)	
								

        therm_data = Landsat_therm_data(Bands,input_folder,Name_Landsat_Image,output_folder,ulx_dem,lry_dem,lrx_dem,uly_dem,shape)          
        Surface_temp=Calc_surface_water_temp(Temp_inst,Landsat_nr,Lmax,Lmin,therm_data,b10_emissivity,k1_c,k2_c,eact_inst,shape,water_mask_temp,Bands_thermal,Rp,tau_sky,surf_temp_offset,Image_Type)
        therm_data_FileName = os.path.join(output_folder,'Surface_Temperature_LS_%s.tif' %Var_name)
        save_GeoTiff_proy(dest, Surface_temp, therm_data_FileName, shape, nband=1)	
								
#################### Calculate NDVI and SAVI for VIIRS-PROBAV ##########################################	
     
    if Image_Type == 2:	
    
        # Define the bands that will be used
        bands=['SM', 'B1', 'B2', 'B3', 'B4']  #'SM', 'BLUE', 'RED', 'NIR', 'SWIR'

        # Set the index number at 0
        index=0
							
        # create a zero array with the shape of the reprojected DEM file						
        data_PROBAV=np.zeros((shape[1], shape[0]))
        spectral_reflectance_PROBAV=np.zeros([shape[1], shape[0], 5])
       
        # constants
        n188_float=248       # Now it is 248, but we do not exactly know what this really means and if this is for constant for all images.
   
        # write the data one by one to the spectral_reflectance_PROBAV       
        for bandnmr in bands:

            # Translate the PROBA-V names to the Landsat band names                
            Band_number={'SM':7,'B1':8,'B2':10,'B3':9,'B4':11}

            # Split the PROBA-V name (Do no change the names but keep the name as downloaded data from VITO)           
            Name_PROBAV=Name_PROBAV_Image.split('_')

            # Get the tile numbers of the PROBA-V           
            X_PROBAV=int(Name_PROBAV[3][1:3])
            Y_PROBAV=int(Name_PROBAV[3][4:6])

            # Define the upperleft coordinate based on the tile number           
            X_ul=-180+(10*X_PROBAV)
            Y_ul=75-(10*Y_PROBAV)
 
            # Open the .hdf file              
            Band_PROBAVhdf_fileName = os.path.join(input_folder, '%s.HDF5' % (Name_PROBAV_Image))   
                    
            # Open the dataset
            g=gdal.Open(Band_PROBAVhdf_fileName)
        
            # open the subdataset to get the projection
            sds_b3 = gdal.Open(g.GetSubDatasets()[Band_number[bandnmr]][0])
            Data = sds_b3.GetRasterBand(1).ReadAsArray()
 
            # Define the x and y spacing           
            X_spacing=float(10.0/int(Data.shape[0]))
            Y_spacing=float(10.0/int(Data.shape[1]))
            
            # Define the georeference of the PROBA-V data
            geo_PROBAV=[X_ul, X_spacing, 0, Y_ul, 0, -Y_spacing]
		 									
            # Define the name of the output file											
            PROBAV_data_name=os.path.join(input_folder, '%s_%s.tif' % (Name_PROBAV_Image,bandnmr)) 									
            dst_fileName=os.path.join(input_folder, PROBAV_data_name)
            
            # create gtiff output with the PROBA-V band
            fmt = 'GTiff'
            driver = gdal.GetDriverByName(fmt)

            dst_dataset = driver.Create(dst_fileName, int(Data.shape[1]), int(Data.shape[0]), 1,gdal.GDT_Float32)
            dst_dataset.SetGeoTransform(geo_PROBAV)
            
            # set the reference info
            srs = osr.SpatialReference()
            srs.SetWellKnownGeogCS("WGS84")
            dst_dataset.SetProjection(srs.ExportToWkt())
           
            # write the array in the geotiff band
            dst_dataset.GetRasterBand(1).WriteArray(Data)
            dst_dataset = None
 
            # Open the PROBA-V band in SEBAL											
            g=gdal.Open(PROBAV_data_name.replace("\\","/")) 
	 										
            # If the data cannot be opened, change the extension											
            if g is None:
                PROBAV_data_name=os.path.join(input_folder, '%s_%s.tiff' % (Name_PROBAV_Image,bandnmr))  
  
            # Reproject the PROBA-V band  to match DEM's resolution          
            PROBAV = reproject_dataset_example(
                              PROBAV_data_name, lon_fileName)

            # Open the reprojected PROBA-V band data                         
            data_PROBAV_DN = PROBAV.GetRasterBand(1).ReadAsArray(0, 0, ncol, nrow)
	 										
            # Define the filename to store the cropped Landsat image
            dst_FileName = os.path.join(output_folder, 'Output_PROBAV','proy_PROBAV_%s.tif' % bandnmr)
		 									
            # close the PROBA-V 											
            g=None
                                   
            # If the band data is not SM change the DN values into PROBA-V values and write into the spectral_reflectance_PROBAV                      
            if bandnmr is not 'SM':  
                data_PROBAV[:, :]=data_PROBAV_DN/2000                           
                spectral_reflectance_PROBAV[:, :, index]=data_PROBAV[:, :]
               
            # If the band data is the SM band than write the data into the spectral_reflectance_PROBAV  and create cloud mask            
            else:
                data_PROBAV[:, :]=data_PROBAV_DN
                Cloud_Mask_PROBAV=np.zeros((shape[1], shape[0]))
                Cloud_Mask_PROBAV[data_PROBAV[:,:]!=n188_float]=1
                spectral_reflectance_PROBAV[:, :, index]=Cloud_Mask_PROBAV
  
            # Change the spectral reflectance to meet certain limits                               
            spectral_reflectance_PROBAV[:, :, index]=np.where(spectral_reflectance_PROBAV[:, :, index]<=0,np.nan,spectral_reflectance_PROBAV[:, :, index])   
            spectral_reflectance_PROBAV[:, :, index]=np.where(spectral_reflectance_PROBAV[:, :, index]>=150,np.nan,spectral_reflectance_PROBAV[:, :, index])   
  										
            # Go to the next index 									
            index=index+1
 
        # Bands in PROBAV spectral reflectance
        # 0 = MS
        # 1 = BLUE
        # 2 = NIR
        # 3 = RED
        # 4 = SWIR
       
        # Calculate SAVI based on PROBA-V
        L = 0.5							
        SAVI = (1+L)*(spectral_reflectance_PROBAV[:, :, 3]-spectral_reflectance_PROBAV[:, :, 2])/(L+spectral_reflectance_PROBAV[:, :, 2]+spectral_reflectance_PROBAV[:, :, 3])
 
        # Calculate surface albedo based on PROBA-V
        Surface_Albedo_PROBAV = 0.219 * spectral_reflectance_PROBAV[:, :, 1] + 0.361 * spectral_reflectance_PROBAV[:, :, 2] + 0.379 * spectral_reflectance_PROBAV[:, :, 3] + 0.041 * spectral_reflectance_PROBAV[:, :, 4]
  
        # Create Water mask based on PROBA-V             
        water_mask_temp = np.zeros((shape[1], shape[0])) 
        water_mask_temp[np.logical_and(spectral_reflectance_PROBAV[:, :, 2] >= spectral_reflectance_PROBAV[:, :, 3],data_DEM>0)]=1

        # Reproject the Veg_height to the LAI projection
        # dest = reproject_dataset3(Veg_Height_proj_FileName, LAI_FileName)			
        Albedo_FileName = os.path.join(output_folder,'Albedo_PROBAV_%s.tif' %Var_name) 
								
        save_GeoTiff_proy(dest, Surface_Albedo_PROBAV, Albedo_FileName, shape, nband=1)	  

        # Calculate the NDVI based on PROBA-V     
        n218_memory = spectral_reflectance_PROBAV[:, :, 2] + spectral_reflectance_PROBAV[:, :, 3]
        NDVI = np.zeros((shape[1], shape[0]))
        NDVI[n218_memory != 0] =  ( spectral_reflectance_PROBAV[:, :, 3][n218_memory != 0] - spectral_reflectance_PROBAV[:, :, 2][n218_memory != 0] )/ ( spectral_reflectance_PROBAV[:, :, 2][n218_memory != 0] + spectral_reflectance_PROBAV[:, :, 3][n218_memory != 0] )

        NDVI_FileName = os.path.join(output_folder,'NDVI_PROBAV_%s.tif' %Var_name)
        save_GeoTiff_proy(dest, NDVI, NDVI_FileName, shape, nband=1)	

        SAVI_FileName = os.path.join(output_folder,'SAVI_PROBAV_%s.tif' %Var_name)
        save_GeoTiff_proy(dest, SAVI, SAVI_FileName, shape, nband=1)															
				
        # Calculate and save the LAI based on NDVI and SAVI 
        FPAR,tir_emis,Nitrogen,vegt_cover,LAI,b10_emissivity=Calc_vegt_para(NDVI,SAVI,water_mask_temp,shape)
        LAI_FileName = os.path.join(output_folder,'LAI_PROBAV_%s.tif' %Var_name) 				
        save_GeoTiff_proy(dest, LAI, LAI_FileName, shape, nband=1)	

################################## Calculate VIIRS surface temperature ########################

        # Define the VIIRS thermal data name
        VIIRS_data_name=os.path.join(input_folder, '%s' % (Name_VIIRS_Image_TB))
							
        # Reproject VIIRS thermal data								
        VIIRS = reproject_dataset_example(VIIRS_data_name, LAI_FileName)
																
        # Open VIIRS thermal data																		
        data_VIIRS = VIIRS.GetRasterBand(1).ReadAsArray()    
				
        # Define the thermal VIIRS output name
        proyVIIRS_fileName = os.path.join(output_folder, 'Surface_Temp_VIIRS_%s.tif' %Var_name)
	 											
        # Save the thermal VIIRS data 												
        save_GeoTiff_proy(dest, data_VIIRS, proyVIIRS_fileName, shape, nband=1)	

################################################### HANTS #######################################################




################################## All input is now calculated, so preprosessing can start ########################

    # Open preprosessing excel the Vegetation_Height sheet				
    ws_veg = wb_veg['Vegetation_Height'] 

    # Define output name for the LandUse map 
    dst_FileName = os.path.join(output_folder,'LU_%s.tif' %Var_name) 

    # Open LU data
    LU_dest = gdal.Open(LU_data_FileName)
    LU_data = LU_dest.GetRasterBand(1).ReadAsArray() 
 
    # Reproject the LAI to the same projection as LU
    dest1 = reproject_dataset_example(LAI_FileName, LU_data_FileName)	 ## input after HANTS
    LAI_proj = dest1.GetRasterBand(1).ReadAsArray() 

    # Read out the excel file coefficient numbers			
    Array = np.zeros([ws_veg.max_row-1,4])
    for j in ['A','C','D','E']:
        j_number={'A' : 0, 'C' : 1, 'D' : 2, 'E' : 3}					
        for i in range(2,ws_veg.max_row+1):											
	        Value = (ws_veg['%s%s' %(j,i)].value)  																
	        Array[i-2, j_number[j]] = Value										

    # Create maps with the coefficient numbers for the right land cover
    coeff = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1]),3])
    for coeff_nmbr in range(0,3):				
        for Class in range(0,len(Array)):
	        coeff[LU_data==Array[Class,0],coeff_nmbr] = Array[Class,coeff_nmbr+1]

    # Get some dimensions of the projected dataset 
    band_data = dest1.GetRasterBand(1) 
    ncol_data = dest1.RasterXSize
    nrow_data = dest1.RasterYSize
    shape_data=[ncol_data, nrow_data]

    # Calculate the vegetation height in the LU projection
    Veg_Height_proj = coeff[:,:,0] * np.power(LAI_proj,2) + coeff[:,:,1] * LAI_proj + coeff[:,:,2]
    Veg_Height_proj = np.clip(Veg_Height_proj, 0, 600)

    # Save the vegetation height in the lU projection in the temporary directory
    Veg_Height_proj_FileName = os.path.join(output_folder_temp,'Veg_Height_proj.tif') 				
    save_GeoTiff_proy(dest1, Veg_Height_proj, Veg_Height_proj_FileName, shape_data, nband=1)	
				
    # Reproject the Veg_height to the LAI projection
    dest = reproject_dataset_example(Veg_Height_proj_FileName, LAI_FileName)			

    # Get some dimensions of the original dataset 
    band_data = dest.GetRasterBand(1)
    ncol_data = dest.RasterXSize
    nrow_data = dest.RasterYSize

    # Open the Veg_height with the same projection as LAI				
    Veg_Height = band_data.ReadAsArray(0, 0, ncol_data, nrow_data)
    Veg_Height[Veg_Height == 0] = np.nan				

    # Save Vegetation Height in the end folder				
    dst_FileName = os.path.join(output_folder,'Vegetation_Height_%s.tif' %Var_name) 	
    save_GeoTiff_proy(dest, Veg_Height, dst_FileName, shape, nband=1)			

######################## calculate p-factor by using the Landuse map #########################
    ws_p = wb_veg['p-factor'] 
			
    Array_P = np.zeros([ws_p.max_row-1,2])
    for j in ['A','C']:
        j_number={'A' : 0, 'C' : 1}					
        for i in range(2,ws_p.max_row+1):											
            Value = (ws_p['%s%s' %(j,i)].value)  																
            Array_P[i-2, j_number[j]] = Value	

    p_factor = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1])])		
    for Class in range(0,len(Array_P)):
	    p_factor[LU_data==Array_P[Class,0]] = Array_P[Class,1]

    p_factor[p_factor == 0] = np.nan

    dst_FileName = os.path.join(output_folder_temp,'p-factor_proj.tif') 	
    save_GeoTiff_proy(dest1, p_factor, dst_FileName, shape_data, nband=1)

    dest = reproject_dataset_example(dst_FileName, LAI_FileName)	

    band_data = dest.GetRasterBand(1) # Get the reprojected dem band	
    ncol_data = dest.RasterXSize
    nrow_data = dest.RasterYSize				
    p_factor = band_data.ReadAsArray(0, 0, ncol_data, nrow_data)
    p_factor[p_factor == 0] = np.nan

    dst_pfactor_FileName = os.path.join(output_folder,'p-factor_%s.tif' %Var_name) 
    save_GeoTiff_proy(dest, p_factor, dst_pfactor_FileName, shape, nband=1)	

######################## calculate c-factor by using the Landuse map #########################

    ws_c = wb_veg['C-factor'] 
			
    Array_C = np.zeros([ws_c.max_row-1,2])
    for j in ['A','C']:
        j_number={'A' : 0, 'C' : 1}					
        for i in range(2,ws_c.max_row+1):											
            Value = (ws_c['%s%s' %(j,i)].value)  																
            Array_C[i-2, j_number[j]] = Value	

    c_factor = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1])])		
    for Class in range(0,len(Array_C)):
	    c_factor[LU_data==Array_C[Class,0]] = Array_C[Class,1]

    c_factor[np.logical_and(c_factor != 3.0, c_factor != 4.0)] = np.nan

    LUE_max = np.zeros([int(np.shape(LU_data)[0]),int(np.shape(LU_data)[1])])	
    LUE_max[c_factor == 3] = 2.5
    LUE_max[c_factor == 4] = 4.5
    LUE_max[LUE_max == 0] = np.nan

    dst_FileName = os.path.join(output_folder_temp,'LUE_max_proj.tif') 	
    save_GeoTiff_proy(dest1, LUE_max, dst_FileName, shape_data, nband=1)

    dest = reproject_dataset_example(dst_FileName, LAI_FileName)	

    band_data = dest.GetRasterBand(1) # Get the reprojected dem band	
    ncol_data = dest.RasterXSize
    nrow_data = dest.RasterYSize				
    LUE_max = band_data.ReadAsArray(0, 0, ncol_data, nrow_data)
    LUE_max[LUE_max == 0] = np.nan

    dst_LUEmax_FileName = os.path.join(output_folder,'LUE_max_%s.tif' %Var_name) 
    save_GeoTiff_proy(dest, LUE_max, dst_LUEmax_FileName, shape, nband=1)	

############################# delete temporary directory ########################
    shutil.rmtree(output_folder_temp)
#################################################################################
# Functions
#################################################################################   
def DEM_lat_lon(DEM_fileName,output_folder_temp):
    """
    This function retrieves information about the latitude and longitude of the
    DEM map. 
    
    """
    # name for output
    lat_fileName = os.path.join(output_folder_temp,'latitude.tif')
    lon_fileName = os.path.join(output_folder_temp,'longitude.tif')
				
				
    g = gdal.Open(DEM_fileName)     # Open DEM
    geo_t = g.GetGeoTransform()     # Get the Geotransform vector:
    x_size = g.RasterXSize          # Raster xsize - Columns
    y_size = g.RasterYSize          # Raster ysize - Rows
    
    # create a longitude and a latitude array 
    lon = np.zeros((y_size, x_size))
    lat = np.zeros((y_size, x_size))
    for col in np.arange(x_size):
        lon[:, col] = geo_t[0] + col * geo_t[1] + geo_t[1]/2
        # ULx + col*(E-W pixel spacing) + E-W pixel spacing
    for row in np.arange(y_size):
        lat[row, :] = geo_t[3] + row * geo_t[5] + geo_t[5]/2
        # ULy + row*(N-S pixel spacing) + N-S pixel spacing,
        # negative as we will be counting from the UL corner

    # Save lat and lon files in geo- coordinates
    save_GeoTiff_geo(g, lat, lat_fileName, x_size, y_size, nband=1)
    save_GeoTiff_geo(g, lon, lon_fileName, x_size, y_size, nband=1)
    
    return(lat,lon,lat_fileName,lon_fileName)

#------------------------------------------------------------------------------
def Calc_Gradient(dataset,pixel_spacing):
    """
    This function calculates the slope and aspect of a DEM map.
    """
    # constants
    deg2rad = np.pi / 180.0  # Factor to transform from degree to rad
    rad2deg = 180.0 / np.pi  # Factor to transform from rad to degree
    
    # calulate slope from DEM map
    x, y = np.gradient(dataset)
    slope = np.arctan(np.sqrt(np.square(x/pixel_spacing) + np.square(y/pixel_spacing))) * rad2deg
    
    # calculate aspect                  
    aspect = np.arctan2(y/pixel_spacing, -x/pixel_spacing) * rad2deg
    aspect = 180 + aspect

    return(deg2rad,rad2deg,slope,aspect)

#------------------------------------------------------------------------------ 
def Calc_Ra_Mountain(lon,DOY,hour,minutes,lon_proy,lat_proy,slope,aspect):    
    """
    Calculates the extraterrestiral solar radiation by using the date, slope and aspect.
    """
    
    # Constants
    deg2rad = np.pi / 180.0  # Factor to transform from degree to rad
    Min_cos_zn = 0.1  # Min value for cos zenith angle
    Max_cos_zn = 1.0  # Max value for cos zenith angle
    Loc_time = float(hour) + float(minutes)/60  # Local time (hours)
    
    # 1. Calculation of extraterrestrial solar radiation for slope and aspect
    # Computation of Hour Angle (HRA = w)
    B = 360./365 * (DOY-81)           # (degrees)
    # Computation of cos(theta), where theta is the solar incidence angle
    # relative to the normal to the land surface
    delta=np.arcsin(np.sin(23.45*deg2rad)*np.sin(np.deg2rad(B))) # Declination angle (radians)
    phi = lat_proy * deg2rad                                     # latitude of the pixel (radians)
    s = slope * deg2rad                                          # Surface slope (radians)
    gamma = (aspect-180) * deg2rad                               # Surface aspect angle (radians)
    w=w_time(Loc_time, lon_proy, DOY)                            # Hour angle (radians)
    a,b,c = Constants(delta,s,gamma,phi)
    cos_zn= AngleSlope(a,b,c,w)
    cos_zn = cos_zn.clip(Min_cos_zn, Max_cos_zn)
    
    dr = 1 + 0.033 * cos(DOY*2*pi/365)  # Inverse relative distance Earth-Sun

    return(dr,cos_zn)

#------------------------------------------------------------------------------
def Constants(delta,s,gamma,phi):
    '''
    Based on Richard G. Allen 2006 equation 11
    determines constants for calculating the exterrestial solar radiation
    '''
    a = np.sin(delta)*np.cos(phi)*np.sin(s)*np.cos(gamma) - np.sin(delta)*np.sin(phi)*np.cos(s)
    b = np.cos(delta)*np.cos(phi)*np.cos(s) + np.cos(delta)*np.sin(phi)*np.sin(s)*np.cos(gamma)
    c = np.cos(delta)*np.sin(s)*np.sin(gamma)
    return(a,b,c)

#------------------------------------------------------------------------------				
def w_time(LT,lon_proy, DOY):
    """
    This function computes the hour angle (radians) of an image given the
    local time, longitude, and day of the year.

    """
    nrow, ncol = lon_proy.shape
   
    # Difference of the local time (LT) from Greenwich Mean Time (GMT) (hours):
    delta_GTM = np.sign(lon_proy[nrow/2, ncol/2]) * lon_proy[nrow/2, ncol/2] * 24 / 360
    if np.isnan(delta_GTM) == True:
         delta_GTM = np.nanmean(lon_proy) * np.nanmean(lon_proy)  * 24 / 360
    
    
    # Local Standard Time Meridian (degrees):
    LSTM = 15 * delta_GTM
    
    # Ecuation of time (EoT, minutes):
    B = 360./365 * (DOY-81)  # (degrees)
    EoT = 9.87*sin(np.deg2rad(2*B))-7.53*cos(np.deg2rad(B))-1.5*sin(np.deg2rad(B))
   
    # Net Time Correction Factor (minutes) at the center of the image:
    TC = 4 * (lon_proy - LSTM) + EoT     # Difference in time over the longitude
    LST = LT + delta_GTM + TC/60         # Local solar time (hours)
    HRA = 15 * (LST-12)                  # Hour angle HRA (degrees)
    deg2rad = np.pi / 180.0              # Factor to transform from degree to rad
    w = HRA * deg2rad                    # Hour angle HRA (radians)
    return w

#------------------------------------------------------------------------------				
def AngleSlope(a,b,c,time):
    '''
    Based on Richard G. Allen 2006
    Calculate the cos zenith angle by using the hour angle and constants
    '''
    angle = -a + b*np.cos(time) + c*np.sin(time)
    return(angle)    

#------------------------------------------------------------------------------
def Landsat_therm_data(Bands,input_folder,Name_Landsat_Image,output_folder,ulx_dem,lry_dem,lrx_dem,uly_dem,shape_lsc):          
    """
    This function calculates and returns the thermal data from the landsat image.
    """                             
    therm_data = np.zeros((shape_lsc[1], shape_lsc[0], len(Bands)-6))
    for band in Bands[-(len(Bands)-6):]:
        # Open original Landsat image for the band number
        src_FileName = os.path.join(input_folder, '%s_B%1d.TIF'
                                    % (Name_Landsat_Image, band))
        if not os.path.exists(src_FileName):
             src_FileName = os.path.join(input_folder, '%s_B%1d_VCID_2.TIF'
                                    % (Name_Landsat_Image, band))		    																																
																																				
        # Define the filename to store the cropped Landsat image
        dst_FileName = os.path.join(output_folder, 'Temp',
                                    'cropped_LS_b%1d.tif' % band)

        ls_data=Open_landsat(src_FileName,dst_FileName,ulx_dem,lry_dem,lrx_dem,uly_dem,shape_lsc) 																																	

        index = np.where(Bands[:] == band)[0][0] - 6
        therm_data[:, :, index] = ls_data
								
    return(therm_data)

#------------------------------------------------------------------------------
def Landsat_Reflect(Bands,input_folder,Name_Landsat_Image,output_folder_temp,ulx_dem,lry_dem,lrx_dem,uly_dem,shape_lsc,Lmax,Lmin,ESUN_L5,ESUN_L7,ESUN_L8,cos_zn_resh,dr,Landsat_nr):
    """
    This function calculates and returns the reflectance and spectral radiation from the landsat image.
    """ 
    
    Reflect = np.zeros((shape_lsc[1], shape_lsc[0], 7))
    for band in Bands[:-(len(Bands)-6)]:
        # Open original Landsat image for the band number
        src_FileName = os.path.join(input_folder, '%s_B%1d.TIF'
                                    % (Name_Landsat_Image, band))
        # Define the filename to store the cropped Landsat image
        dst_FileName = os.path.join(output_folder_temp,
                                    'cropped_LS_b%1d.tif' % band)
          
        ls_data=Open_landsat(src_FileName,dst_FileName,ulx_dem,lry_dem,lrx_dem,uly_dem,shape_lsc,gdalwarp)
        # stats = band_data.GetStatistics(0, 1)

        index = np.where(Bands[:-(len(Bands)-6)] == band)[0][0]
        if Landsat_nr == 8:
            # Spectral radiance for each band:
            L_lambda = Landsat_L_lambda(Lmin,Lmax,ls_data,index,Landsat_nr)
            # Reflectivity for each band:
            rho_lambda = Landsat_rho_lambda(L_lambda,ESUN_L8,index,cos_zn_resh,dr)
        elif Landsat_nr == 7:
            # Spectral radiance for each band:
            L_lambda=Landsat_L_lambda(Lmin,Lmax,ls_data,index,Landsat_nr)
            # Reflectivity for each band:
            rho_lambda = Landsat_rho_lambda(L_lambda,ESUN_L7,index,cos_zn_resh,dr)
        elif Landsat_nr == 5:
            # Spectral radiance for each band:
            L_lambda=Landsat_L_lambda(Lmin,Lmax,ls_data,index,Landsat_nr)
            # Reflectivity for each band:
            rho_lambda =Landsat_rho_lambda(L_lambda,ESUN_L5,index,cos_zn_resh,dr)
        else:
            print 'Landsat image not supported, use Landsat 5, 7 or 8'

        Reflect[:, :, index] = rho_lambda
    Reflect = Reflect.clip(0.0, 1.0)
    return(Reflect)

#------------------------------------------------------------------------------				
def Open_landsat(src_FileName,dst_FileName,ulx_dem,lry_dem,lrx_dem,uly_dem,shape_lsc):
    """
    This function opens a landsat image and returns the data array of a specific landsat band.
    """                           
    # crop band to the DEM extent
    fullCmd = ' '.join(['gdalwarp -te %s %s %s %s' % (ulx_dem, lry_dem,lrx_dem, uly_dem), src_FileName, dst_FileName])
    process = subprocess.Popen(fullCmd)
    process.wait()
    
    # Open the cropped Landsat image for the band number
    ls = gdal.Open(dst_FileName)
    band_data = ls.GetRasterBand(1)
    ls_data = band_data.ReadAsArray(0, 0, shape_lsc[0], shape_lsc[1])
    return(ls_data) 

#------------------------------------------------------------------------------				
def Landsat_L_lambda(Lmin,Lmax,ls_data,index,Landsat_nr):
    """
    Calculates the lambda from landsat
    """
    if Landsat_nr==8:
        L_lambda = ((Lmax[index] - Lmin[index]) / (65535 - 1) * ls_data + Lmin[index]) 
    elif Landsat_nr == 5 or Landsat_nr ==7:
        L_lambda = (Lmax[index] - Lmin[index]) / 255 * ls_data + Lmin[index]
    return(L_lambda)

#------------------------------------------------------------------------------				
def Landsat_rho_lambda(L_lambda,ESUN,index,cos_zn_resh,dr):
    """
    Calculates the lambda from landsat
    """
    rho_lambda = np.pi * L_lambda / (ESUN[index] * cos_zn_resh * dr)
    return(rho_lambda)
   				
#------------------------------------------------------------------------------
def Calc_NDVI_SAVI_albedo(Reflect):
    """
    This function calculates and returns the Surface albedo, NDVI, and SAVI by using the refectance from the landsat image.
    """
    # Computation of Normalized Difference Vegetation Index (NDVI)
    # and Soil Adjusted Vegetation Index (SAVI):
    L = 0.5				
    NDVI = ((Reflect[:, :, 3] - Reflect[:, :, 2]) /
            (Reflect[:, :, 3] + Reflect[:, :, 2]))
    SAVI = (1 + L) * ((Reflect[:, :, 3] - Reflect[:, :, 2]) /
                      (L + Reflect[:, :, 3] + Reflect[:, :, 2]))

    Apparent_atmosf_transm = 0.89    # This value is used for atmospheric correction of broad band albedo. This value is used for now, would be better to use tsw.
    path_radiance = 0.03             # Recommended, Range: [0.025 - 0.04], based on Bastiaanssen (2000).
																						
    # Surface albedo:
    Surf_albedo = (0.254 * Reflect[:, :, 0] + 0.149 * Reflect[:, :, 1] +
                   0.147 * Reflect[:, :, 2] + 0.311 * Reflect[:, :, 3] +
                   0.103 * Reflect[:, :, 4] + 0.036 * Reflect[:, :, 5] -
                   path_radiance) / np.power(Apparent_atmosf_transm, 2)

    # Better tsw instead of Apparent_atmosf_transm ??
    Surf_albedo = Surf_albedo.clip(0.0, 0.6)																						
																						
    return(NDVI,SAVI,Surf_albedo)

#------------------------------------------------------------------------------
def Calc_surface_water_temp(Temp_inst,Landsat_nr,Lmax,Lmin,therm_data,b10_emissivity,k1_c,k2_c,eact,shape_lsc,water_mask_temp,Bands_thermal,Rp,tau_sky,surf_temp_offset,Image_Type):    
    """
    Calculates the surface temperature and create a water mask
    """ 
    # Spectral radiance for termal
    if Landsat_nr == 8:
        if Bands_thermal == 1:
            k1 = k1_c[0]
            k2 = k2_c[0]
            L_lambda_b10 = (Lmax[-1] - Lmin[-1]) / (65535-1) * therm_data[:, :, 0] + Lmin[-1]
            
            # Get Temperature
            Temp_TOA = Get_Thermal(L_lambda_b10,Rp,Temp_inst,tau_sky,b10_emissivity,k1,k2) 
                              
        elif Bands_thermal == 2:
            L_lambda_b10 = (Lmax[-2] - Lmin[-2]) / (65535-1) * therm_data[:, :, 0] + Lmin[-2]
            L_lambda_b11 = (Lmax[-1] - Lmin[-1]) / (65535-1) * therm_data[:, :, 1] + Lmin[-1]
    
            # Brightness temperature
            # From Band 10:
            Temp_TOA_10 = (k2_c[0] / np.log(k1_c[0] / L_lambda_b10 + 1.0))
            # From Band 11:
            Temp_TOA_11 = (k2_c[1] / np.log(k1_c[1] / L_lambda_b11 + 1.0))
            # Combined:
            Temp_TOA = (Temp_TOA_10 + 1.378 * (Temp_TOA_10 - Temp_TOA_11) +
                           0.183 * np.power(Temp_TOA_10 - Temp_TOA_11, 2) - 0.268 +
                           (54.30 - 2.238 * eact) * (1 - b10_emissivity))
       
    elif Landsat_nr == 7:
        k1=666.09
        k2=1282.71
        L_lambda_b6 = (Lmax[-1] - Lmin[-1]) / (256-1) * therm_data[:, :, 0] + Lmin[-1]

        # Brightness temperature - From Band 6:
        Temp_TOA = Get_Thermal(L_lambda_b6,Rp,Temp_inst,tau_sky,b10_emissivity,k1,k2) 
        
    elif Landsat_nr == 5:
        k1=607.76
        k2=1260.56
        L_lambda_b6 = ((Lmax[-1] - Lmin[-1]) / (256-1) * therm_data[:, :, 0] +
                       Lmin[-1])
   
       # Brightness temperature - From Band 6:
        Temp_TOA = Get_Thermal(L_lambda_b6,Rp,Temp_inst,tau_sky,b10_emissivity,k1,k2) 
            
    # Surface temperature
    Surface_temp = Temp_TOA
    Surface_temp = Surface_temp.clip(230.0, 360.0)

    return(Surface_temp)    

#------------------------------------------------------------------------------
def Get_Thermal(lambda_b10,Rp,Temp_inst,tau_sky,TIR_Emissivity,k1,k2):
    
    # Narrow band downward thermal radiation from clear sky, rsky (W/m2/sr/µm)
    rsky = (1.807E-10 * np.power(Temp_inst + 273.15, 4) * (1 - 0.26 *
            np.exp(-7.77E-4 * np.power((-Temp_inst), -2))))
    print 'Rsky = ', '%0.3f (W/m2/sr/µm)' % np.nanmean(rsky)
    
    # Corrected thermal radiance from the surface, Wukelikc et al. (1989):
    correc_lambda_b10 = ((lambda_b10 - Rp) / tau_sky -
                               (1.0 - TIR_Emissivity) * rsky)
    # Brightness temperature - From Band 10:
    Temp_TOA = (k2 / np.log(TIR_Emissivity * k1 /
                       correc_lambda_b10 + 1.0))  
                       
    return(Temp_TOA)           

#------------------------------------------------------------------------------        
def Calc_vegt_para(NDVI,SAVI,water_mask_temp,shape_lsc):
    """
    Calculates the Fraction of PAR, Thermal infrared emissivity, Nitrogen, Vegetation Cover, LAI, b10_emissivity
    """
    # Fraction of PAR absorbed by the vegetation canopy (FPAR):
    FPAR = -0.161 + 1.257 * NDVI
    FPAR[NDVI < 0.125] = 0.0

    # Termal infrared emissivity
    tir_emis = 1.009 + 0.047 * np.log(NDVI)
    tir_emis[np.logical_or(water_mask_temp == 1.0, water_mask_temp == 2.0)] = 1.0
    tir_emis[np.logical_and(NDVI < 0.125, water_mask_temp == 0.0)] = 0.92

    # Vegetation Index - Regression model from Bagheri et al. (2013)
    VI_NDVI = 38.764 * np.square(NDVI) - 24.605 * NDVI + 5.8103
    VI_SAVI = 6.3707 * np.square(SAVI) - 2.8503 * SAVI + 1.6335
    VI = (VI_NDVI + VI_SAVI) / 2.0  # Average of computed from NDVI and SAVI

    # Nitrogen computation
    Nitrogen = np.copy(VI)
    Nitrogen[VI <= 0.0] = 0.0
    Nitrogen[NDVI <= 0.0] = 0.0
                  
    # Vegetation cover:
    vegt_cover = 1 - np.power((0.8 - NDVI)/(0.8 - 0.125), 0.7)
    vegt_cover[NDVI < 0.125] = 0.0
    vegt_cover[NDVI > 0.8] = 0.99

    # Leaf Area Index (LAI)
    LAI_1 = np.log(-(vegt_cover - 1)) / -0.45
    LAI_1[LAI_1 > 8] = 8.0
    LAI_2 = (9.519 * np.power(NDVI, 3) + 0.104 * np.power(NDVI, 2) +
             1.236 * NDVI - 0.257)
    LAI_3 = 11.0 * np.power(SAVI, 3)
    LAI_3[SAVI >= 0.817] = 6.0
    LAI_4 = -np.log((0.69 - SAVI) / 0.59) / 0.91  # For South. Idaho, empirical
    LAI_4[SAVI < 0.0] = 0.0
    LAI_4[SAVI >= 0.689] = 6.0

    LAI = (LAI_1 + LAI_2 + LAI_3 + LAI_4) / 4.0  # Average LAI
    LAI[LAI < 0.001] = 0.001

    b10_emissivity = np.zeros((shape_lsc[1], shape_lsc[0]))
    b10_emissivity = np.where(LAI <= 3.0, 0.95 + 0.01 * LAI, 0.98)
    b10_emissivity[water_mask_temp != 0.0] = 1.0
    return(FPAR,tir_emis,Nitrogen,vegt_cover,LAI,b10_emissivity)

#------------------------------------------------------------------------------
def info_band_metadata(filename, Bands):
    """
    This function retrieves Landsat band information (minimum and maximum
    radiance) from the metadata file.

    """
    Lmin = np.zeros(len(Bands))  # Minimum band radiance, for each band
    Lmax = np.zeros(len(Bands))  # Maximum band radiance, for each band
    k1_const = np.zeros(len(Bands)-6)  # TIRS_Thermal constant k1 ######
    k2_const = np.zeros(len(Bands)-6)  # TIRS_Thermal constant k2 ######
    for band in Bands:
        Landsat_meta = open(filename, "r")  # Open metadata file
        for line in Landsat_meta:
            if re.match("(.*)RADIANCE_MINIMUM_BAND_%1d(.*)" % band, line):
                words = line.split()
                value = float(words[2])
                Lmin[np.where(Bands == band)[0][0]] = value
            if re.match("(.*)RADIANCE_MAXIMUM_BAND_%1d(.*)" % band, line):
                words = line.split()
                value = float(words[2])
                Lmax[np.where(Bands == band)[0][0]] = value
            if re.match("(.*)K1_CONSTANT_BAND_%1d(.*)" % band, line):  # #####
                words = line.split()
                value = float(words[2])
                k1_const[np.where(Bands == band)[0][0]-6] = value
            if re.match("(.*)K2_CONSTANT_BAND_%1d(.*)" % band, line):  # #####
                words = line.split()
                value = float(words[2])
                k2_const[np.where(Bands == band)[0][0]-6] = value
    return Lmin, Lmax, k1_const, k2_const

#------------------------------------------------------------------------------				
def info_general_metadata(filename):
    """
    This function retrieves general information of the Landsat image
    (date and time aquired, UTM zone, sun elevation) from the
    metadata file.

    """
    Landsat_meta = open(filename, "r")  # Open metadata file
    for line in Landsat_meta:
        if re.match("(.*)SCENE_CENTER_TIME(.*)", line): # search in metadata for line SCENE_CENTER_TIME
            words = line.split()# make groups of words which are divided by an open space
            time_list = words[2].split(':', 2) # Take the second word of words and split the word which are divided by :
            if len(time_list[0])== 3: 
                time_list[0]=time_list[0][1:3]
                time_list[2]=time_list[2][0:-1]
            hour = float(time_list[0]) # take the first word of time_list
            minutes = float(time_list[1]) + float(time_list[2][:-1]) / 60 # Take the second and third word of time_list and place :-1 to remove Z behined minutes
    Landsat_meta = open(filename, "r")  # Open metadata file    
    for line in Landsat_meta:
        if re.match("(.*)DATE_ACQUIRED(.*)", line):
            words = line.split()
            DOY = time.strptime(words[2], "%Y-%m-%d").tm_yday
            year = time.strptime(words[2], "%Y-%m-%d").tm_year
    Landsat_meta = open(filename, "r")  # Open metadata file
    for line in Landsat_meta:
        if re.match("(.*)UTM_ZONE(.*)", line):
            words = line.split()
            UTM_Zone = int(words[2])
    Landsat_meta = open(filename, "r")  # Open metadata file
    for line in Landsat_meta:
        if re.match("(.*)SUN_ELEVATION(.*)", line):
            words = line.split()
            Sun_elevation = float(words[2])

    return year, DOY, hour, minutes, UTM_Zone, Sun_elevation

#------------------------------------------------------------------------------				
def reproject_dataset(dataset, pixel_spacing, UTM_Zone):
    """
    A sample function to reproject and resample a GDAL dataset from within
    Python. The idea here is to reproject from one system to another, as well
    as to change the pixel size. The procedure is slightly long-winded, but
    goes like this:

    1. Set up the two Spatial Reference systems.
    2. Open the original dataset, and get the geotransform
    3. Calculate bounds of new geotransform by projecting the UL corners
    4. Calculate the number of pixels with the new projection & spacing
    5. Create an in-memory raster dataset
    6. Perform the projection
    """

    # 1) Open the dataset
    g = gdal.Open(dataset)
    if g is None:
        print 'input folder does not exist'

    epsg_from = Get_epsg(g)	
   
    # Get the Geotransform vector:
    geo_t = g.GetGeoTransform()
    # Vector components:
    # 0- The Upper Left easting coordinate (i.e., horizontal)
    # 1- The E-W pixel spacing
    # 2- The rotation (0 degrees if image is "North Up")
    # 3- The Upper left northing coordinate (i.e., vertical)
    # 4- The rotation (0 degrees)
    # 5- The N-S pixel spacing, negative as it is counted from the UL corner
    x_size = g.RasterXSize  # Raster xsize
    y_size = g.RasterYSize  # Raster ysize
    DEM_UL_lat = geo_t[3]

    # Define the EPSG code...
    if DEM_UL_lat > 0:
        EPSG_code = '326%02d' % UTM_Zone
    else:
        EPSG_code = '326%02d' % UTM_Zone
        UTM_Zone = - UTM_Zone
    epsg_to = int(EPSG_code)

    # 2) Define the UK OSNG, see <http://spatialreference.org/ref/epsg/27700/>
    osng = osr.SpatialReference()
    osng.ImportFromEPSG(epsg_to)
    wgs84 = osr.SpatialReference()
    wgs84.ImportFromEPSG(epsg_from)

    inProj = Proj(init='epsg:%d' %epsg_from)
    outProj = Proj(init='epsg:%d' %epsg_to)
				
    # Up to here, all  the projection have been defined, as well as a
    # transformation from the from to the to

    # 3) Work out the boundaries of the new dataset in the target projection
    #   Skip some rows and columns in the border to avoid null values due to
    #   reprojection - rectangle to parallelogram
    nrow_skip = round((0.10*y_size)/2)
    ncol_skip = round((0.10*x_size)/2)
    

    ulx, uly = transform(inProj,outProj,geo_t[0]+ncol_skip*geo_t[1], geo_t[3] +
                       nrow_skip * geo_t[5])
    lrx, lry = transform(inProj,outProj,geo_t[0] + geo_t[1] * (x_size-ncol_skip),
                                        geo_t[3] + geo_t[5] * (y_size-nrow_skip))

    # See how using 27700 and WGS84 introduces a z-value!
    # Now, we create an in-memory raster
    mem_drv = gdal.GetDriverByName('MEM')

    # The size of the raster is given the new projection and pixel spacing
    # Using the values we calculated above. Also, setting it to store one band
    # and to use Float32 data type.
    col = int((lrx - ulx)/pixel_spacing)
    rows = int((uly - lry)/pixel_spacing)

    # Re-define lr coordinates based on whole number or rows and columns
    (ulx, uly) = (int(ulx), int(uly))
    (lrx, lry) = (int(ulx) + col * pixel_spacing, int(uly) -
                  rows * pixel_spacing)
    dest = mem_drv.Create('', col, rows, 1, gdal.GDT_Float32)
    if dest is None:
        print 'input folder to large for memory, clip input map'
     
   # Calculate the new geotransform
    new_geo = (int(ulx), pixel_spacing, geo_t[2], int(uly),
               geo_t[4], - pixel_spacing)
    
    # Set the geotransform
    dest.SetGeoTransform(new_geo)
    dest.SetProjection(osng.ExportToWkt())
      
    # Perform the projection/resampling
    res = gdal.ReprojectImage(g, dest, wgs84.ExportToWkt(), osng.ExportToWkt(),gdal.GRA_Bilinear)
				
    return dest, ulx, lry, lrx, uly, epsg_to
				
#------------------------------------------------------------------------------
def reproject_dataset_example(dataset, dataset_example):

    # open dataset that must be transformed    
    g = gdal.Open(dataset)
    epsg_from = Get_epsg(g)	   

    # open dataset that is used for transforming the dataset
    gland=gdal.Open(dataset_example) 
    epsg_to = Get_epsg(gland)	

    # Set the EPSG codes
    osng = osr.SpatialReference()
    osng.ImportFromEPSG(epsg_to)
    wgs84 = osr.SpatialReference()
    wgs84.ImportFromEPSG(epsg_from)

    # Get shape and geo transform from example				
    geo_land = gland.GetGeoTransform()			
    col=gland.RasterXSize
    rows=gland.RasterYSize

    # Create new raster			
    mem_drv = gdal.GetDriverByName('MEM')
    dest1 = mem_drv.Create('', col, rows, 1, gdal.GDT_Float32)
    dest1.SetGeoTransform(geo_land)
    dest1.SetProjection(osng.ExportToWkt())
    
    # Perform the projection/resampling
    res = gdal.ReprojectImage(g, dest1, wgs84.ExportToWkt(), osng.ExportToWkt(), gdal.GRA_NearestNeighbour)

    return(dest1)		
				
#------------------------------------------------------------------------------
def Get_epsg(g):				
			
    try:
        # Get info of the dataset that is used for transforming     
        gland_proj = g.GetProjection()
        Projection=gland_proj.split('EPSG","')
        epsg_to=int((str(Projection[-1]).split(']')[0])[0:-1])				      
    except:
        epsg_to=4326	
        print 'Was not able to get the projection, so WGS84 is assumed'							
    return(epsg_to)	
				
#------------------------------------------------------------------------------				
def Get_Extend_Landsat(src_FileName):
    """
    This function gets the extend of the landsat image
    """
    ls = gdal.Open(src_FileName)       # Open Landsat image
    print 'Original LANDSAT Image - '
    geo_t_ls = ls.GetGeoTransform()    # Get the Geotransform vector
    x_size_ls = ls.RasterXSize         # Raster xsize - Columns
    y_size_ls = ls.RasterYSize         # Raster ysize - Rows
    print '  Size :', x_size_ls, y_size_ls
    (ulx, uly) = geo_t_ls[0], geo_t_ls[3]
    (lrx, lry) = (geo_t_ls[0] + geo_t_ls[1] * x_size_ls,
                  geo_t_ls[3] + geo_t_ls[5] * y_size_ls)
    band_data = ls.GetRasterBand(1)
    
    return(ls,band_data,ulx,uly,lrx,lry,x_size_ls,y_size_ls)
    				
#------------------------------------------------------------------------------
def save_GeoTiff_geo(src_dataset, dst_dataset_array, dst_fileName, ncol, nrow,
                     nband):
    """
    This function saves an array dataset in GeoTiff, using the parameters
    from the source dataset, in geographical coordinates

    """
    geotransform = src_dataset.GetGeoTransform()
    # create dataset for output
    fmt = 'GTiff'
    driver = gdal.GetDriverByName(fmt)
    dir_name = os.path.dirname(dst_fileName)
    # If the directory does not exist, make it.
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)
    dst_dataset = driver.Create(dst_fileName, ncol, nrow, nband,gdal.GDT_Float32)
    dst_dataset.SetGeoTransform(geotransform)
    # set the reference info
    srs = osr.SpatialReference()
    srs.SetWellKnownGeogCS("WGS84")
    dst_dataset.SetProjection(srs.ExportToWkt())
    # write the array in the geotiff band
    dst_dataset.GetRasterBand(1).WriteArray(dst_dataset_array)
    # stats = dst_dataset.GetRasterBand(1).GetStatistics(0, 1)
    dst_dataset = None				
				
#------------------------------------------------------------------------------
def reshape(src_FileName, dst_FileName, LS_resol, var):
    """
    This function resamples the DEM related maps (lat, lon, etc.)
    to the resolution of the Landsat images (generally 30 m)
    radiance) from the metadata file.

    """
    dir_name = os.path.dirname(dst_FileName)
    
    # If the directory does not exist, make it.
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)
    fullCmd = ' '.join(['gdalwarp -tr %s %s ' % (LS_resol, LS_resol),src_FileName, dst_FileName])  # -r {nearest}
    process = subprocess.Popen(fullCmd)
    process.wait()
    re_var = gdal.Open(dst_FileName)  # Open cropped Image
    
    print '---'
    print 'Reshaping %s - ' % var
    x_size = re_var.RasterXSize     # Raster xsize - Number of Columns
    y_size = re_var.RasterYSize     # Raster ysize - Number of Rows
    band_data_var = re_var.GetRasterBand(1)
    var_resh = band_data_var.ReadAsArray(0, 0, x_size, y_size)
    return var_resh

#------------------------------------------------------------------------------
def Reshape_Reproject_Input_data(input_File_Name, output_File_Name, Example_extend_fileName):
       
   data_rep, ulx_dem, lry_dem, lrx_dem, uly_dem, epsg_to = reproject_dataset_example(
       input_File_Name, Example_extend_fileName)
   band_data = data_rep.GetRasterBand(1) # Get the reprojected dem band
   ncol_data = data_rep.RasterXSize
   nrow_data = data_rep.RasterYSize
   shape_data=[ncol_data, nrow_data]
   
   #stats = band.GetStatistics(0, 1)
   data = band_data.ReadAsArray(0, 0, ncol_data, nrow_data)
   save_GeoTiff_proy(data_rep, data, output_File_Name, shape_data, nband=1)
   return(data)

#------------------------------------------------------------------------------				
def save_GeoTiff_proy(src_dataset, dst_dataset_array, dst_fileName, shape_lsc,nband):
    """
    This function saves an array dataset in GeoTiff, using the parameters
    from the source dataset, in projected coordinates

    """
    geotransform = src_dataset.GetGeoTransform()
    spatialreference = src_dataset.GetProjection()
    # create dataset for output
    fmt = 'GTiff'
    driver = gdal.GetDriverByName(fmt)
    dir_name = os.path.dirname(dst_fileName)
    # If the directory does not exist, make it.
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)
    dst_dataset = driver.Create(dst_fileName, shape_lsc[0], shape_lsc[1], nband,gdal.GDT_Float32)
    dst_dataset.SetGeoTransform(geotransform)
    dst_dataset.SetProjection(spatialreference)
    dst_dataset.GetRasterBand(1).WriteArray(dst_dataset_array)
    dst_dataset = None

#------------------------------------------------------------------------------
def Water_Mask(shape_lsc,Reflect):
    """
    Calculates the water mask
    """
    mask = np.zeros((shape_lsc[1], shape_lsc[0]))
    mask[np.logical_and(Reflect[:, :, 3] < Reflect[:, :, 2],
                        Reflect[:, :, 4] < Reflect[:, :, 1])] = 1.0
    water_mask_temp = np.copy(mask)
    return(water_mask_temp)
				
#------------------------------------------------------------------------------				
if __name__ == '__main__':
    main()